import datetime

import telebot
from telebot import types

import random

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import os


token = "6773724804:AAFgufpFjuSRJyspNu_ONmk3qST_wiXO3D4"

bot = telebot.TeleBot(token)

file_path = os.path.abspath('C:/Users/Kikita/OneDrive/Документы/bot_schedule_maker/Расписание_тренировок.xlsx')

excel_send_package = []

gyms = ['Зал 1', 'Зал 2']

trainers = ['Лера', 'Ксюша', 'Яна', 'Таня', 'Юля', 'Лена', 'Настя', 'Настя Е.']

week = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

time = ['10:00-11:00', '11:00-12:00', '12:00-13:00', '13:00-14:00', '16:00-17:00', '17:00-18:00', '18:00-19:00',
        '19:00-20:00', '20:00-21:00']

work_type = ['ПРО ягодицы', 'Здоровая спина', 'Скайстретчинг', 'Классическая растяжка', 'Растяжка с подкачкой', 
             'Пилатес', 'Йога', 'Йога в гамаках', 'ТРХ', 'Аэростретчинг', 'Аэродети', 'МФР', 'Детская растяжка',
             'Ролл-стретчинг']


# Преобразовываем числовой месяц в буквенный
def transform_date():
    month, year = [datetime.datetime.now().month, datetime.datetime.now().year]
    months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
              'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    return f'{months[int(month)]} {year}'


# Преобразовываем числовой номер столбца ячейки в буквенный
def transform_coordinate(a):
    letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    return letter[a - 1]


# Получение координаты ячейки
def get_coordinate():
    wb = load_workbook(file_path)

    sheet = wb.get_sheet_by_name(transform_date())  # Поменять на функцию transform_date по наступлению января 2024

    if excel_send_package[0] == 'Зал 1':
        for c_column in sheet['A2':'H13']:
            for cell in c_column:
                if cell.value == excel_send_package[1]:
                    column_num = cell.column
        for c_row in sheet['A2':'H13']:
            for cell in c_row:
                if cell.value == excel_send_package[2]:
                    row_num = cell.row

    if excel_send_package[0] == 'Зал 2':
        for c_column in sheet['A16':'H27']:
            for cell in c_column:
                if cell.value == excel_send_package[1]:
                    column_num = cell.column              
        for c_row in sheet['A16':'H27']:
            for cell in c_row:
                if cell.value == excel_send_package[2]:
                    row_num = cell.row

    wb.close()

    return [transform_coordinate(column_num) + str(row_num)]


# Записываем значение в ячейку
def write_cell(a, b):
    wb = load_workbook(file_path)

    sheet = wb.get_sheet_by_name(transform_date())  # Поменять на функцию transform_date по наступлению января 2024

    sheet[str(*get_coordinate())] = a + ' - ' + b

    if a == trainers[0]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="D2691E")
    elif a == trainers[1]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="FFDEAD")
    elif a == trainers[2]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="228B22")
    elif a == trainers[3]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="4169E1")
    elif a == trainers[4]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="B0C4DE")
    elif a == trainers[5]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="EE82EE")
    elif a == trainers[6]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="9ACD32")
    elif a == trainers[7]:
        sheet[str(*get_coordinate())].fill = PatternFill('solid', fgColor="DAA520")

    wb.save(file_path)
    wb.close


# Клавиатура со списком залов
def inline_gym():
    ingymmarkup = types.InlineKeyboardMarkup(row_width=1)
    btn1 = types.InlineKeyboardButton(gyms[0], callback_data=gyms[0])
    btn2 = types.InlineKeyboardButton(gyms[1], callback_data=gyms[1])
    ingymmarkup.add(btn1, btn2)
    return ingymmarkup


# Клавиатура с тренерами
def inline_trainers():
    inMurkup = types.InlineKeyboardMarkup(row_width=2)
    item1 = types.InlineKeyboardButton(trainers[0], callback_data=trainers[0])
    item2 = types.InlineKeyboardButton(trainers[1], callback_data=trainers[1])
    item3 = types.InlineKeyboardButton(trainers[2], callback_data=trainers[2])
    item4 = types.InlineKeyboardButton(trainers[3], callback_data=trainers[3])
    item5 = types.InlineKeyboardButton(trainers[4], callback_data=trainers[4])
    item6 = types.InlineKeyboardButton(trainers[5], callback_data=trainers[5])
    item7 = types.InlineKeyboardButton(trainers[6], callback_data=trainers[6])
    item8 = types.InlineKeyboardButton(trainers[7], callback_data=trainers[7])
    inMurkup.add(item1, item2, item3, item4, item5, item6, item7, item8)
    return inMurkup


# Клавиатура с днями недели
def inline_week():
    inMarkup = types.InlineKeyboardMarkup(row_width=2)
    but1 = types.InlineKeyboardButton(week[0], callback_data=week[0])
    but2 = types.InlineKeyboardButton(week[1], callback_data=week[1])
    but3 = types.InlineKeyboardButton(week[2], callback_data=week[2])
    but4 = types.InlineKeyboardButton(week[3], callback_data=week[3])
    but5 = types.InlineKeyboardButton(week[4], callback_data=week[4])
    but6 = types.InlineKeyboardButton(week[5], callback_data=week[5])
    but7 = types.InlineKeyboardButton(week[6], callback_data=week[6])
    inMarkup.add(but1, but2, but3, but4, but5, but6, but7)
    return inMarkup


# Клавиатура со временем
def inline_time():
    murkup = types.InlineKeyboardMarkup(row_width=2)
    but1 = types.InlineKeyboardButton(time[0], callback_data=time[0])
    but2 = types.InlineKeyboardButton(time[1], callback_data=time[1])
    but3 = types.InlineKeyboardButton(time[2], callback_data=time[2])
    but4 = types.InlineKeyboardButton(time[3], callback_data=time[3])
    but5 = types.InlineKeyboardButton(time[4], callback_data=time[4])
    but6 = types.InlineKeyboardButton(time[5], callback_data=time[5])
    but7 = types.InlineKeyboardButton(time[6], callback_data=time[6])
    but8 = types.InlineKeyboardButton(time[7], callback_data=time[7])
    but9 = types.InlineKeyboardButton(time[8], callback_data=time[8])
    murkup.add(but1, but2, but3, but4, but5, but6, but7, but8, but9)
    return murkup


# Клавиатура с тренировками
def inline_work_type():
    work_type_keyboard = types.InlineKeyboardMarkup(row_width=2)
    but1 = types.InlineKeyboardButton(work_type[0], callback_data=work_type[0])
    but2 = types.InlineKeyboardButton(work_type[1], callback_data=work_type[1])
    but3 = types.InlineKeyboardButton(work_type[2], callback_data=work_type[2])
    but4 = types.InlineKeyboardButton(work_type[3], callback_data=work_type[3])
    but5 = types.InlineKeyboardButton(work_type[4], callback_data=work_type[4])
    but6 = types.InlineKeyboardButton(work_type[5], callback_data=work_type[5])
    but7 = types.InlineKeyboardButton(work_type[6], callback_data=work_type[6])
    but8 = types.InlineKeyboardButton(work_type[7], callback_data=work_type[7])
    but9 = types.InlineKeyboardButton(work_type[8], callback_data=work_type[8])
    but10 = types.InlineKeyboardButton(work_type[9], callback_data=work_type[9])
    but11 = types.InlineKeyboardButton(work_type[10], callback_data=work_type[10])
    but12 = types.InlineKeyboardButton(work_type[11], callback_data=work_type[11])
    but13 = types.InlineKeyboardButton(work_type[12], callback_data=work_type[12])
    but14 = types.InlineKeyboardButton(work_type[13], callback_data=work_type[13])
    work_type_keyboard.add(but1, but2, but3, but4, but5, but6, but7, but8, but9, but10, but11, but12, but13, but14)
    return work_type_keyboard


# Клавиатура с кнопками завершения работы
def inline_back_end_keyboard():
    back_end_keyboard = types.InlineKeyboardMarkup(row_width=2)
    back_btn = types.InlineKeyboardButton('Вернуться к выбору тренера', callback_data='Вернуться к выбору тренера')
    end_btn = types.InlineKeyboardButton('Выгрузить файл', callback_data='Выгрузить файл')
    back_end_keyboard.add(back_btn, end_btn)
    return back_end_keyboard


# Отправка первого сообщения и создание первой клавиатуры
@bot.message_handler(commands=['start'])
def start_message(message):
    stic = ['sticker.webm', 'sticker1.webm', 'sticker2.webm']
    send_stic = open('C:/Users/Kikita/OneDrive/Документы/bot_schedule_maker/' + random.choice(stic), 'rb')

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    item1 = types.KeyboardButton("Ага...(((")
    item2 = types.KeyboardButton("Пока нет))) Просто выгрузи мне файл с расписанием")
    markup.add(item1)
    markup.add(item2)
    
    bot.send_message(message.chat.id, "Привет, Валерка, снова пора делать расписание?", reply_markup=markup)
    bot.send_sticker(message.chat.id, send_stic)


# Скачивание файла расписания
@bot.message_handler(func=lambda message: True)
def download_schedule(message):
    doc = open(file_path, 'rb')

    if message.chat.type == 'private':
        if message.text == "Пока нет))) Просто выгрузи мне файл с расписанием":
            bot.send_message(message.chat.id, 'Вот твое расписание')
            bot.send_document(message.chat.id, doc)
        else:
            bot.send_message(message.chat.id, "Я не знаю что и ответить")


# Создание клавиатуры со списком залов
@bot.message_handler(func=lambda message: True)
def get_gym(message):
    if message.chat.type == 'private':
        if message.text == "Ага...(((":
            bot.send_message(message.chat.id, 'Выбери пыточную', reply_markup=inline_gym())
        elif message.text == "Жека струковое поле":
            send_egg_stic = open('C:/Users/Kikita/OneDrive/Документы/bot_schedule_maker/stickereasteregg.webm', 'rb')
            bot.send_sticker(message.chat.id, send_egg_stic)
        else:
            bot.send_message(message.chat.id, "Я не знаю что и ответить")


# Создание клавиатууры с днями недели
@bot.callback_query_handler(func=lambda call: call.data in gyms)
def get_day(call):
    try:  
        if call.message:
            if call.data in gyms:
                excel_send_package.append(call.data)
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, 
                                      text='Отлично, теперь выбери день недели', reply_markup=inline_week())
            
    except Exception as e:
        print(repr(e))


# Создание клавиатуры со временем
@bot.callback_query_handler(func=lambda call: call.data in week)
def get_time(call):
    try:  
        if call.message:
            if call.data in week:
                excel_send_package.append(call.data)
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, 
                                      text='Пора выбрать время', reply_markup=inline_time())
            bot.answer_callback_query(callback_query_id=call.id, show_alert=False)

    except Exception as e:
        print(repr(e))  


# Создание клавиатуры с тренерами
@bot.callback_query_handler(func=lambda call: call.data in time)
def get_trainer(call):
    try:  
        if call.message:
            if call.data in time:
                excel_send_package.append(call.data)
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, 
                                      text='Choose your fighter', reply_markup=inline_trainers())
            bot.answer_callback_query(callback_query_id=call.id, show_alert=False)

    except Exception as e:
        print(repr(e))  


# Создание клавиатуры с типом тренировки
@bot.callback_query_handler(func=lambda call: call.data in trainers)
def get_work_type(call):
    try:  
        if call.message:
            if call.data in trainers:
                excel_send_package.append(call.data)
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, 
                                      text='Выбери метод экзекуции', reply_markup=inline_work_type())
            bot.answer_callback_query(callback_query_id=call.id, show_alert=False)
    except Exception as e:
        print(repr(e))


# Создание клавиатуры с кнопками завершения
@bot.callback_query_handler(func=lambda call: call.data in work_type)
def get_back_end(call):
    try:
        if call.message:
            excel_send_package.append(call.data)
            if call.data in work_type:
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, 
                                      text='Чаво делаим дальше?', reply_markup=inline_back_end_keyboard())
                bot.send_message(call.message.chat.id, 'Ты выбрала:' + str(excel_send_package[:]))
                bot.answer_callback_query(callback_query_id=call.id, show_alert=False)
    except Exception as e:
        print(repr(e))


# Запись значения в ячейку и продолжение работы с ботом, посредством перехода к клавиатуре со списком залов
@bot.callback_query_handler(func=lambda call: call.data == 'Вернуться к выбору тренера')
def get_back_end(call):
    try:
        if call.message:
            if call.data == 'Вернуться к выбору тренера': 
                write_cell(excel_send_package[3], excel_send_package[4])
                excel_send_package.clear()
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, 
                                      text='Выбери пыточную', reply_markup=inline_gym())
                bot.answer_callback_query(callback_query_id=call.id, show_alert=False)
    except Exception as e:
        print(repr(e))


# Запись значения в ячейку и завершение работы с ботом, выгрузка файла в чат
@bot.callback_query_handler(func=lambda call: call.data == 'Выгрузить файл')
def get_back_end(call):
    try:
        if call.message:
            if call.data == 'Выгрузить файл':
                write_cell(excel_send_package[3], excel_send_package[4])  
                doc = open(file_path, 'rb')
                bot.answer_callback_query(callback_query_id=call.id, text='Файл выгружается', show_alert=True) 
                bot.send_message(call.message.chat.id, 'Вот твое расписание')
                bot.send_document(call.message.chat.id, doc)
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text='То, что ты выбирала:', reply_markup=None)
             
    except Exception as e:
        print(repr(e))


bot.polling(none_stop=True)
